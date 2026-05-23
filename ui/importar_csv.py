import asyncio
import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import flet as ft
from dateutil import parser as dateutil_parser

from entities.ternero import Ternero
from printings.impresion_terneros import imprimir_terneros
from savings import savings
from savings.leer_directorio import autocargar_imagenes_ternero, validar_carpeta_imagenes

_TABLE_HEADERS = ["Nombre", "R.P.", "R.Control", "Raza", "HBA",
                  "Fecha de Nacimiento", "Foto Izq.", "Foto Der."]

_SUPPORTED_EXTENSIONS = ["csv", "xlsx", "xlsm", "xls"]


# ── File reading ───────────────────────────────────────────────────────────────

def _read_file(path: str) -> list[list[Any]]:
    """Reads CSV or Excel file and returns raw rows (values may be str/datetime/int/float)."""
    ext = Path(path).suffix.lower().lstrip(".")
    if ext == "csv":
        return _read_csv(path)
    elif ext in ("xlsx", "xlsm"):
        return _read_xlsx(path)
    elif ext == "xls":
        return _read_xls(path)
    else:
        raise ValueError(f"Formato no soportado: .{ext}")


def _read_csv(path: str) -> list[list[str]]:
    with open(path, "rb") as f:
        raw = f.read()
    delimiter = ";" if raw.count(b";") > raw.count(b",") else ","
    text = raw.decode("utf-8", errors="replace")
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def _read_xlsx(path: str) -> list[list[Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _read_xls(path: str) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError:
        raise ImportError(
            "Para leer archivos .xls instalá xlrd:\n  pip install xlrd==1.2.0"
        )
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    rows: list[list[Any]] = []
    for r in range(ws.nrows):
        row: list[Any] = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
            else:
                row.append(cell.value)
        rows.append(row)
    return rows


# ── Value coercions ────────────────────────────────────────────────────────────

def _str(val: Any) -> str:
    if val is None:
        return ""
    # openpyxl returns floats for numeric-looking cells (e.g. RPs like 123.0)
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _date(val: Any) -> Optional[date]:
    """Parse a date from any value: datetime object, or string in any common format."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan", "-", ""):
        return None
    try:
        # dayfirst=True → prefers DD/MM/YYYY over MM/DD/YYYY (Argentine convention)
        return dateutil_parser.parse(s, dayfirst=True).date()
    except (ValueError, OverflowError, TypeError):
        return None


# ── Page class ─────────────────────────────────────────────────────────────────

class ImportPage:
    def __init__(self, page: ft.Page):
        self.page = page
        self.terneros: list[Ternero] = []
        self.carpeta_imagenes: str = ""

    def build(self) -> ft.Control:
        self.folder_picker = ft.FilePicker()
        self.file_picker = ft.FilePicker()

        instructions = ft.Text(
            "Columnas esperadas (fila 1 = encabezado, se ignora):\n"
            "Nombre · RP · RControl · Raza · HBA · Sexo · Color · FechaNac · "
            "PadreRP · PadreHBA · NombrePadre · MadreRP · MadreHBA · MadreCAL · "
            "NombreMadre · FotoIzquierda · FotoDerecha\n"
            f"Formatos aceptados: {', '.join('.' + e for e in _SUPPORTED_EXTENSIONS)}  "
            "— Fechas: cualquier formato reconocible (DD/MM/AAAA, AAAA-MM-DD, etc.)",
            size=12,
            color=ft.Colors.GREY_700,
        )

        self.label_carpeta = ft.Text("No se ha seleccionado la carpeta de imágenes", size=13)

        btn_carpeta = ft.FilledButton(
            "Seleccionar carpeta de imágenes",
            icon=ft.Icons.FOLDER_OPEN,
            on_click=self._on_pick_folder,
        )
        btn_archivo = ft.FilledButton(
            "Seleccionar planilla (CSV / Excel)",
            icon=ft.Icons.UPLOAD_FILE,
            on_click=self._on_pick_file,
        )

        folder_section = ft.Column([
            ft.Text("Carpeta de Imágenes (opcional):", weight=ft.FontWeight.W_500),
            ft.Row([btn_carpeta, self.label_carpeta],
                   vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Text(
                "Formato de nombre: RP + I/D + extensión  (ej: 001I.jpg, 001D.png)",
                size=12,
                color=ft.Colors.GREY_600,
            ),
        ])

        self.data_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text(h, weight=ft.FontWeight.BOLD))
                for h in _TABLE_HEADERS
            ],
            rows=[],
            column_spacing=16,
            horizontal_margin=8,
            heading_row_height=40,
            data_row_min_height=32,
            data_row_max_height=40,
        )

        btn_guardar = ft.FilledButton("Guardar Todos",  icon=ft.Icons.SAVE,      on_click=self._on_guardar_todos)
        btn_imprimir = ft.FilledButton("Imprimir Todos", icon=ft.Icons.PRINT,     on_click=self._on_imprimir_todos)
        btn_limpiar  = ft.OutlinedButton("Limpiar Tabla", icon=ft.Icons.CLEAR_ALL, on_click=self._on_limpiar)

        return ft.Column(
            expand=True,
            spacing=10,
            controls=[
                ft.Column([instructions, folder_section, btn_archivo], spacing=8),
                ft.Divider(),
                ft.Column(controls=[self.data_table], scroll=ft.ScrollMode.AUTO, expand=True),
                ft.Divider(),
                ft.Row([btn_guardar, btn_imprimir, btn_limpiar]),
            ],
        )

    # ── Picker handlers ────────────────────────────────────────────────

    async def _on_pick_folder(self, e: ft.ControlEvent) -> None:
        directory = await self.folder_picker.get_directory_path()
        if not directory:
            return

        self.carpeta_imagenes = directory
        try:
            n = await asyncio.to_thread(validar_carpeta_imagenes, self.carpeta_imagenes)
            self.label_carpeta.value = f"✓ {n} imágenes encontradas"
        except Exception as ex:
            _show_info(self.page, "Error", f"Error en la carpeta: {ex}")
            self.carpeta_imagenes = ""
            self.label_carpeta.value = "Error al leer la carpeta"
            self.page.update()
            return

        self.page.update()
        if self.terneros:
            await asyncio.to_thread(self._autocargar_imagenes)
            self._refresh_table()

    async def _on_pick_file(self, e: ft.ControlEvent) -> None:
        files = await self.file_picker.pick_files(
            allow_multiple=False,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=_SUPPORTED_EXTENSIONS,
        )
        if not files:
            return

        path = files[0].path
        try:
            records = await asyncio.to_thread(_read_file, path)
        except Exception as ex:
            _show_info(self.page, "Error al leer el archivo", str(ex))
            return

        if len(records) < 2:
            _show_info(self.page, "Aviso", "El archivo está vacío o sólo tiene encabezado")
            return

        self.terneros = []
        fechas_fallidas = 0

        for i, row in enumerate(records):
            if i == 0:
                continue  # skip header
            if len(row) < 17:
                continue  # skip incomplete rows

            fecha = _date(row[7])
            if row[7] and not fecha:
                fechas_fallidas += 1

            self.terneros.append(Ternero(
                nombre=_str(row[0]),
                rp=_str(row[1]),
                r_de_control=_str(row[2]),
                raza=_str(row[3]),
                hba=_str(row[4]),
                sexo=_str(row[5]),
                color=_str(row[6]),
                fecha_nac=fecha,
                padre_rp=_str(row[8]),
                padre_hba=_str(row[9]),
                nombre_padre=_str(row[10]),
                madre_rp=_str(row[11]),
                madre_hba=_str(row[12]),
                madre_cal=_str(row[13]),
                nombre_madre=_str(row[14]),
                foto_izquierda=_str(row[15]),
                foto_derecha=_str(row[16]),
            ))

        if not self.terneros:
            _show_info(self.page, "Aviso", "No se encontraron filas válidas")
            return

        if self.carpeta_imagenes:
            await asyncio.to_thread(self._autocargar_imagenes)
        else:
            msg = f"Se cargaron {len(self.terneros)} terneros."
            if fechas_fallidas:
                msg += f"\n⚠ {fechas_fallidas} fechas no pudieron interpretarse."
            msg += "\nSeleccione una carpeta de imágenes para cargarlas automáticamente."
            _show_info(self.page, "Aviso", msg)

        self._refresh_table()

        extra = f"  ({fechas_fallidas} fechas sin reconocer)" if fechas_fallidas else ""
        _show_info(self.page, "Éxito", f"Se cargaron {len(self.terneros)} terneros{extra}")

    # ── Action handlers ────────────────────────────────────────────────

    async def _on_guardar_todos(self, e: ft.ControlEvent) -> None:
        if not self.terneros:
            _show_info(self.page, "Aviso", "No hay datos para guardar")
            return
        exitosos, errores = await asyncio.to_thread(savings.guardar_multiple, self.terneros)
        if errores:
            _show_info(self.page, "Completado con errores",
                       f"Se generaron {exitosos} de {len(self.terneros)} planillas. {errores} errores.")
        else:
            _show_info(self.page, "Éxito",
                       f"Se generaron {exitosos} planillas en la carpeta 'planillas_lote'")

    async def _on_imprimir_todos(self, e: ft.ControlEvent) -> None:
        if not self.terneros:
            _show_info(self.page, "Aviso", "No hay datos para imprimir")
            return

        progress_dlg = ft.AlertDialog(
            title=ft.Text("Generando planillas..."),
            content=ft.Column(
                [ft.ProgressRing(), ft.Text("Por favor espere...")],
                tight=True,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            modal=True,
        )
        self.page.show_dialog(progress_dlg)

        error_msg: str | None = None
        try:
            await asyncio.to_thread(imprimir_terneros, list(self.terneros))
        except Exception as ex:
            error_msg = str(ex)

        self.page.pop_dialog()
        if error_msg:
            _show_info(self.page, "Error al imprimir", error_msg)

    def _on_limpiar(self, e: ft.ControlEvent) -> None:
        self.terneros = []
        self._refresh_table()

    # ── Helpers ───────────────────────────────────────────────────────

    def _autocargar_imagenes(self) -> None:
        if not self.carpeta_imagenes:
            return
        cargadas = 0
        for t in self.terneros:
            izq, der = autocargar_imagenes_ternero(t.rp, self.carpeta_imagenes)
            if izq:
                t.foto_izquierda = izq
                cargadas += 1
            if der:
                t.foto_derecha = der
                cargadas += 1
        msg = (f"Se cargaron automáticamente {cargadas} imágenes"
               if cargadas
               else "No se encontraron imágenes que coincidan con los RP de los terneros")
        _show_info(self.page, "Imágenes" if cargadas else "Aviso", msg)

    def _refresh_table(self) -> None:
        self.data_table.rows.clear()
        for t in self.terneros:
            fecha_str = t.fecha_nac.strftime("%d/%m/%Y") if t.fecha_nac else "-"
            self.data_table.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(t.nombre)),
                ft.DataCell(ft.Text(t.rp)),
                ft.DataCell(ft.Text(t.r_de_control)),
                ft.DataCell(ft.Text(t.raza)),
                ft.DataCell(ft.Text(t.hba)),
                ft.DataCell(ft.Text(fecha_str)),
                ft.DataCell(ft.Text("✓" if t.foto_izquierda else "✗")),
                ft.DataCell(ft.Text("✓" if t.foto_derecha else "✗")),
            ]))
        self.page.update()


# ── Module helper ──────────────────────────────────────────────────────────────

def _show_info(page: ft.Page, title: str, message: str) -> None:
    def on_ok(e):
        page.pop_dialog()
    page.show_dialog(ft.AlertDialog(
        title=ft.Text(title),
        content=ft.Text(message),
        actions=[ft.TextButton("OK", on_click=on_ok)],
    ))
